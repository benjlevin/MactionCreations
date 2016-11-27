import openpyxl
import time
import os
import threading
import random
from openpyxl.styles import Font, Alignment, NamedStyle, Side, Border, PatternFill
from flask import send_from_directory

# List to store temporary directories and scheduler to clear list at next midnight after function call
temp_dirs = []
last_clear_time = 0.0


def create_workbook():
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = 'Rank List'

    b1 = ws['B1']
    d1 = ws['D1']

    b1.value = 'Applicant A Rank List'
    b1.alignment = Alignment(horizontal='center')
    b1.font = Font(size=12)
    ws.merge_cells('B1:C1')

    d1.value = 'Applicant B Rank List'
    d1.alignment = Alignment(horizontal='center')
    d1.font = Font(size=12)
    ws.merge_cells('D1:E1')

    for row in ws['D2':'D400']:
        for cell in row:
            cell.border += Border(left=Side(style='thin'))
    for row in ws['F2':'F400']:
        for cell in row:
            cell.border += Border(left=Side(style='thin'))

    b2 = ws['B2']
    c2 = ws['C2']
    d2 = ws['D2']
    e2 = ws['E2']
    f2 = ws['F2']
    g2 = ws['G2']
    headings = [b2, c2, d2, e2, f2, g2]
    column_heading = NamedStyle(name="column_heading")
    column_heading.font = Font(bold=True)
    bottom_border = Side(style='medium', color='000000')
    top_border = Side(style='thin')
    column_heading.border = Border(top=top_border, bottom=bottom_border)
    b2.value = 'Rank'
    d2.value = 'Rank'
    c2.value = 'Program'
    e2.value = 'Program'
    f2.value = 'Average Rank'
    g2.value = 'Distance (mi)'
    for heading in headings:
        heading.style = column_heading
        heading.alignment = Alignment(horizontal='center')

    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 5
    ws.column_dimensions['E'].width = 55
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    return workbook


def populate_workbook(parsed_json, workbook):
    ws = workbook.active

    rank_count = 0
    for rank in parsed_json:
        rank_count += 1
        target_row = str(rank_count + 2)

        program_a = rank['programA']
        program_b = rank['programB']
        average = rank['averageRank']
        distance = rank['distance']

        ws['B' + target_row].value = rank_count
        ws['D' + target_row].value = rank_count
        ws['C' + target_row].value = program_a
        ws['E' + target_row].value = program_b
        ws['F' + target_row].value = average
        ws['G' + target_row].value = distance

        ws['B' + target_row].alignment = Alignment(horizontal='center', vertical='center')
        ws['C' + target_row].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['D' + target_row].alignment = Alignment(horizontal='center', vertical='center')
        ws['E' + target_row].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws['F' + target_row].alignment = Alignment(horizontal='center', vertical='center')
        ws['G' + target_row].alignment = Alignment(horizontal='center', vertical='center')

        ws['G' + target_row].number_format = '0.0'

        if rank['exceedsLimit']:
            ws['B' + target_row].fill = PatternFill('solid', fgColor='d16262')
            ws['D' + target_row].fill = PatternFill('solid', fgColor='d16262')
            ws['C' + target_row].fill = PatternFill('solid', fgColor='d16262')
            ws['E' + target_row].fill = PatternFill('solid', fgColor='d16262')
            ws['F' + target_row].fill = PatternFill('solid', fgColor='d16262')
            ws['G' + target_row].fill = PatternFill('solid', fgColor='d16262')

    return workbook


def create_xls(input_data):
    workbook = create_workbook()
    workbook = populate_workbook(input_data, workbook)
    return workbook


def handle_xml(request):
    dir_path = os.path.dirname(os.path.abspath(__file__))
    if request.method == 'POST':
        if request.headers['Content-Type'] == 'application/json':
            rank_list_data = request.get_json()
            if not rank_list_data:
                return 'N'

            filename = str(random.randint(1e12, 1e13 - 1))
            match_xls = create_xls(rank_list_data)
            match_xls.save(os.path.join(dir_path, 'user_content', 'rank_lists', filename + '.xlsx'))
            return 'Y' + filename
    elif request.method == 'GET':
        filename = request.args.get('f')
        print('Filename: ' + str(filename))
        return send_from_directory(os.path.join(dir_path, 'user_content', 'rank_lists'),
                                   filename=filename + '.xlsx',
                                   as_attachment=True,
                                   attachment_filename='RankList.xlsx')


def couples_cleanup():
    print('Cleaning up couples rank lists.')
    dir_path = os.path.dirname(os.path.abspath(__file__))
    rank_lists = os.path.join(dir_path, 'user_content', 'rank_lists')
    all_lists = []
    for (dirpath, dirnames, filenames) in os.walk(rank_lists):
        all_lists.extend(filenames)
        break
    for file in all_lists:
        file_path = os.path.join(rank_lists, file)
        last_modified = os.stat(file_path).st_mtime
        if time.time() - last_modified > 86400:
            print('Deleting ' + file_path)
            os.remove(file_path)


threading.Timer(86400, couples_cleanup).start()
